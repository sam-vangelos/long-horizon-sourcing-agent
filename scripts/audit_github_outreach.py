#!/usr/bin/env python3

from __future__ import annotations

import argparse
import html
import json
import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROLE_PITCH = (
    "a role centered on data quality, eval design, and frontier-model "
    "improvement work with AI lab clients"
)

STOPWORDS = {
    "and",
    "benchmark",
    "data",
    "evaluation",
    "framework",
    "github",
    "hook",
    "instincts",
    "leadership",
    "llm",
    "message",
    "notes",
    "pipeline",
    "project",
    "repo",
    "role",
    "systems",
    "the",
    "toolkit",
    "work",
}


ELEVATED_SUBJECT_OVERRIDES = {
    "Matheus Augusto": "Your ML platform + model-building background",
    "Miguel Arquez Abdala": "Your production ML engineering background",
    "Jonathan Cristovão": "Question about your ML/data-quality work",
    "Harol Estevez": "Your AI systems work and a data-quality role",
    "Cristina Gomez": "Your NLP/speech ML background",
    "Maria Camila Escobar": "Your CV/data-quality intuition",
    "Peter J. Liu": "Question about your data-centric ML background",
    "Cristina González": "Your multimodal research background",
    "Mike Fuller": "Question about your AI/ML work",
    "Summer-Enzhi": "Your benchmark/evaluation-oriented ML work",
    "NeuralMind": "Who led your NLP data/eval work?",
    "IBM - LLM LAB": "Who led your LLM evaluation work?",
    "H.IAAC": "Who led CAPIVARA/FairPIVARA?",
}


ELEVATED_MESSAGE_OVERRIDES = {
    "Ruan Chaves": (
        "Napolab stood out to me less as a Portuguese benchmark and more as evidence that you take evaluation seriously as a design problem: "
        "curated tasks, a public leaderboard, and a healthy skepticism about what benchmarks do and do not really measure. "
        "We are hiring for a role that lives in that exact gap between model-improvement goals and rigorous evaluation design: datasets, rubrics, validators, "
        "and quality systems for frontier labs. Thought there could be real overlap if you would be open to a brief conversation."
    ),
    "Bryan Oliveira": (
        "What stood out in sliding-puzzles-gym was not just the RL angle, but the way the benchmark forces representation quality to be examined rather than assumed. "
        "Paired with the RL-to-LLM training work on your profile, it reads like the work of someone who cares about the integrity of the training signal, not just the final metric. "
        "We are hiring for a judgment-heavy data/eval role for frontier labs, and I would love to compare notes if that sounds interesting."
    ),
    "Daniel Hoyos": (
        "Your Temporal Fusion Transformer reimplementation reads like the work of someone who respects the difference between getting a model to run and reproducing it faithfully. "
        "That kind of methodological care and attention to model behavior is very close to the work we are hiring for: turning frontier-lab goals into datasets, checks, and evaluation scaffolding that can actually be trusted. "
        "If that kind of hands-on, judgment-heavy role is at all interesting, I would be glad to share more."
    ),
    "Carlo di Francescantonio": (
        "The part of your RL work that stood out to me was the environment design itself: custom Gymnasium worlds, simultaneous action spaces, and the bridge between simulation and training pipelines. "
        "That usually signals someone who thinks carefully about what an agent is really being asked to do and how success gets measured, which is unusually relevant to the role we are filling. "
        "We are hiring for hands-on data/eval work for frontier labs, and I would love to compare notes if that sounds interesting."
    ),
    "Alejandro Aristizábal": (
        "What I found interesting in your work was not just that it touches alignment, but that it treats failure modes and reward misspecification as first-class objects of study rather than afterthoughts. "
        "That mindset — being precise about where an objective can go wrong, and building the evaluation machinery around it — maps very closely to the role we are filling. "
        "If a judgment-heavy data/eval role tied to frontier-model work is of interest, I would be happy to share more."
    ),
    "Jordi Neil Sánchez": (
        "mcp-databricks-server stood out because it sits in a part of the stack most people wave away: the messy interface between LLMs, tools, data systems, and real operational state. "
        "That usually requires strong instincts about schemas, failure handling, and what good tool-mediated behavior actually looks like, which is very close to the quality-system work we are hiring for. "
        "If that overlap sounds real to you, I would love to compare notes."
    ),
    "Vinícius Trevisan": (
        "What I liked about claude-code-supervisor is that it is not just another agent wrapper; it is really about deciding when an output is good enough to keep and when the loop needs another pass. "
        "That kind of judgment about quality thresholds, failure recovery, and iterative improvement is unusually close to the role we are filling for frontier-lab data/eval work. "
        "If you would be open to it, I would love to share a bit more."
    ),
    "Matheus Augusto": (
        "Your public profile reads like someone who has spent real time at the intersection of ML platform work and applied model-building, which is often where the best instincts about data quality get formed. "
        "We are hiring for a role that is less about shipping product features and more about turning model objectives into concrete datasets, validators, and review loops for frontier labs. "
        "I may be extrapolating a bit from the public surface, so if the data/eval side of your work runs deeper than GitHub makes obvious, I would love to compare notes."
    ),
    "Frederico S. Oliveira": (
        "FreeSVC caught my eye not because the domain is singing voice conversion per se, but because the work reflects a real feel for multilingual data complexity, training methodology, and evaluation discipline. "
        "That combination of taste around data and rigor around measurement is very close to what we need in the role we are filling for frontier-model data/eval work. "
        "If the domain shift into LLM/post-training problems is interesting to you, I would love to talk."
    ),
    "Igor Varejão": (
        "What stood out in vibdata was not simply the toolkit itself, but the fact that you are treating data handling and benchmarking as part of model quality rather than as plumbing around it. "
        "That way of thinking — that curation, structure, and evaluation choices quietly shape downstream behavior — is very close to the job we are hiring for. "
        "The domain is different, but the underlying judgment around data quality feels highly transferable if you would be open to a conversation."
    ),
    "Diego Parra": (
        "The interesting part of your fine-tuning work is the full loop: generate the data, define what good output should look like, and then tune and evaluate against that bar. "
        "That is much closer to real data-quality work than a typical fine-tuning project, because the hard part is often setting the right target in the first place. "
        "We are hiring for exactly that kind of judgment-heavy work for frontier labs, and I would be happy to share more if useful."
    ),
    "Miguel Arquez Abdala": (
        "What stood out from your profile was the combination of production ML engineering and strong pipeline discipline rather than any single flashy public repo. "
        "We are hiring for a role where that kind of systematic thinking gets applied to frontier-model data quality: turning fuzzy research needs into concrete datasets, checks, and review loops. "
        "If the data/eval side of your work is a bigger part of your background than GitHub makes obvious, I would love to compare notes."
    ),
    "Gilber A. Corrales": (
        "What I liked about XAI_ARENA is that it treats evaluation as something to formalize, not something to mention after the fact. "
        "Building explicit comparison frameworks and metrics for model behavior is a strong signal for the kind of judgment-heavy quality work we are hiring for. "
        "Different domain, obviously, but the instinct to define what good looks like and then measure against it is exactly the overlap that made me reach out."
    ),
    "Felipe Russi": (
        "Your Cochrane plain-language-summary work stood out because it lives in a hard regime for LLMs: long-form factual transformation where quality depends on structure, faithfulness, and editorial judgment rather than just fluent output. "
        "That is much closer to the substance of the role we are hiring for than a generic agent project would be. "
        "We are filling a hands-on data/eval role for frontier labs, and I would love to compare notes if that sounds interesting."
    ),
    "Jonathan Cristovão": (
        "Your public GitHub is lighter on detailed repo documentation, so I do not want to pretend I have reconstructed the full arc of your work from it alone. "
        "But the profile signal I do see suggests a background in ML/AI work where dataset composition and model behavior matter materially, which is exactly the seam this role lives in. "
        "If that judgment-heavy, data/eval side of the work is interesting to you, I would be glad to share more."
    ),
    "Harol Estevez": (
        "The interesting thread in your work is less RAG in the generic sense and more the way you have had to operationalize quality for domain-specific assistants: retrieval quality, structured outputs, and whether the system is actually useful in context. "
        "That kind of hands-on judgment about what good behavior looks like is very close to the role we are filling for frontier-model data/eval work. "
        "I may be missing some nuance from the public surface, but if there is a deeper quality/evaluation layer to what you have built, I would love to compare notes."
    ),
    "Kevin Scaccia": (
        "Your M4 transformer benchmarking work reads like someone who cares about comparative methodology, not just model selection as a leaderboard exercise. "
        "That matters a lot to us, because the role we are hiring for is really about turning ambiguous model-improvement goals into evaluation logic, data specs, and quality systems that hold up. "
        "If that kind of judgment-heavy ML work is appealing, I would be glad to share more."
    ),
    "Shuhei Kishi": (
        "What stood out in dcraft was not just the tooling; it was the underlying idea of moving data through explicit quality states rather than treating datasets as static artifacts. "
        "That way of thinking maps surprisingly well to the work we are hiring for, where the hard part is often formalizing what raw, trusted, and ready-for-training should actually mean. "
        "If a hands-on role at that intersection of data quality, evaluation, and model improvement is interesting, I would love to compare notes."
    ),
    "Cristina Gomez": (
        "Your public GitHub does not give a lot of repo-level detail, so I do not want to overstate what I can infer from it. "
        "But the profile itself suggests meaningful depth across NLP, speech, and voice systems, which usually goes hand in hand with strong intuitions about annotation quality, data curation, and where model behavior breaks. "
        "We are hiring for a role built around exactly that kind of judgment, and I would be happy to share more if useful."
    ),
    "Maria Camila Escobar": (
        "From the public surface, what stands out is less a single repo and more the kind of work you have orbited: embodied/CV research where the quality of labels, ground truth, and evaluation pipelines does enormous downstream work. "
        "That way of thinking about data quality as part of the science itself is very close to the role we are hiring for. "
        "If the idea of applying that judgment in a more hands-on frontier-data context is interesting, I would love to compare notes."
    ),
    "Carlos D. Escobar-Valbuena": (
        "aiOS stood out because it reflects a very particular kind of systems thinking: define the primitives, make the interfaces explicit, and design around failure rather than assuming the happy path. "
        "We are hiring for a role where that exact mindset gets applied to data and evaluation rather than agent-runtime architecture. "
        "If that crossover sounds interesting, I would be happy to share more."
    ),
    "Nathan Lambert": (
        "What I liked about RLHF Book is that it does not just document methods; it tries to make a messy post-training landscape legible. "
        "That instinct — imposing clarity on an area where people often hand-wave the details — is very close to the work we are hiring for around datasets, validators, and evals for frontier labs. "
        "I realize this may be below your level, but if the data-centric angle is interesting, I would genuinely enjoy comparing notes."
    ),
    "Pengcheng Yin": (
        "What stands out in TranX/NL2code is the depth of thought around what it means for generated code to be not just plausible, but structurally and semantically correct. "
        "That is exactly the kind of judgment frontier teams need when they are building and evaluating coding data, which is why I reached out. "
        "If a hands-on role focused on data quality and evaluation for frontier-model work is at all interesting, I would be glad to share more."
    ),
    "Peter J. Liu": (
        "Your public GitHub is fairly thin, so I do not want to pretend I am grounding this in a rich set of repos. "
        "What did stand out is the profile signal of someone whose work spans serious language-modeling research, which is exactly the kind of background where strong instincts about data quality and post-training judgment usually emerge. "
        "If the idea of applying that judgment in a more hands-on data/eval role is interesting, I would love to compare notes."
    ),
    "Cristina González": (
        "Your public GitHub does not expose much repo-level detail, but the profile signal around CINFONIA and multimodal/vision-language work is enough to make the direction clear. "
        "The reason I am reaching out is that the role we are filling is fundamentally about data judgment: knowing what makes a training example, eval, or annotation scheme genuinely useful rather than merely available. "
        "If that overlap resonates, I would be happy to share more."
    ),
    "Mike Fuller": (
        "Your GitHub surface is fairly light on detailed public projects, so I do not want to overclaim from it. "
        "But between the AI background and the applied data science signal, you look like someone who may already think about model behavior through the lens of data quality rather than just experimentation or deployment. "
        "That is exactly the seam this role sits in, and I would be glad to share more if useful."
    ),
    "Juan Pineda-Jaramillo": (
        "What I liked about llm_query_transport is that it points toward a very practical concern: how to make LLM outputs structured, queryable, and dependable once they touch real systems. "
        "That is adjacent to the kind of quality thinking we are hiring for, where the hard part is often turning vague model goals into precise schemas, checks, and evaluation criteria. "
        "If that kind of work is interesting to you, I would love to compare notes."
    ),
    "Juan Sebastián Corredor Rodriguez": (
        "The interesting part of your image-captioning project is that you built a feedback loop around output quality instead of treating generation as the end of the story. "
        "That instinct — generation plus structured evaluation — is much closer to the substance of our role than a generic training project would be. "
        "We are hiring for hands-on frontier-data/eval work, and I would be glad to share more if it sounds relevant."
    ),
    "Jefferson Hernández": (
        "What stood out in your public work was less any single model result and more the bias toward clean interfaces and validation-minded engineering. "
        "Reimplementing algorithms as usable libraries and thinking in terms of schemas and enforcement is a strong signal for the quality-system side of the role we are filling. "
        "If applying that mindset to frontier-model datasets and evals sounds interesting, I would love to compare notes."
    ),
    "Juan Gutierrez": (
        "Pineapple stood out because paper-reading assistants only become useful when someone has good taste about what to extract, what to discard, and what counts as a faithful summary of the source. "
        "That kind of judgment about useful versus merely fluent model output is very close to the work we are hiring for. "
        "If a hands-on role centered on data quality and evaluation for frontier labs is interesting, I would be happy to share more."
    ),
    "sourabh2k15": (
        "Your training-infrastructure work reads like someone who has spent time close enough to the metal to see how seemingly minor methodological choices cascade into model outcomes. "
        "That perspective is surprisingly relevant to the role we are filling, because the hard part is often diagnosing when a quality issue is really a data issue, an evaluation issue, or a pipeline issue. "
        "If that kind of judgment-heavy work sounds interesting, I would love to compare notes."
    ),
    "David Dinucu-Jianu": (
        "What stood out in actors was not just that it is RL for LLMs, but that it is opinionated about multi-turn structure, environment design, and the practical constraints that shape post-training systems in the real world. "
        "That is very close to the work we are hiring for, where the job is to turn research goals into concrete tasks, data definitions, and quality loops that actually hold up. "
        "If that sounds interesting, I would love to compare notes."
    ),
    "Gustavo Führ": (
        "miyagi_pytorch_trainer stood out because it suggests a methodology-first way of working: make the training stack composable, make the metrics explicit, and make experimentation systematic. "
        "That mindset transfers better to frontier-data work than people often expect, because the real job is often designing the process that keeps quality from drifting. "
        "If the jump from CV into more data/eval-centric model work is interesting, I would be glad to share more."
    ),
    "Summer-Enzhi": (
        "I may be missing some of the exact project surface here, so I do not want to overstate the connection to a single repo. "
        "What did stand out is a broader pattern of benchmark- and evaluation-oriented ML work, which is the real reason I am reaching out. "
        "We are hiring for a role that lives in that zone — turning model goals into datasets, rubrics, validators, and review loops — and I would love to compare notes if that sounds relevant."
    ),
    "NeuralMind": (
        "I am reaching out because NeuralMind's public work suggests deep experience in Portuguese-language NLP where data choices and evaluation methodology are clearly not an afterthought. "
        "We are hiring for a role that sits in that same part of the stack: translating model-improvement goals into datasets, rubrics, validators, and quality systems for frontier labs. "
        "If there is a researcher or engineer on your team who led that work and might be open to a conversation, I would love an introduction."
    ),
    "IBM - LLM LAB": (
        "I am reaching out because the lab's public work suggests real depth in benchmark and evaluation infrastructure for LLMs, which is exactly the part of the stack we are hiring into. "
        "The role is hands-on and centered on turning research goals into datasets, annotation logic, validators, and review loops for frontier-model work. "
        "If there is a specific researcher or engineer behind that evaluation work who might be open to a conversation, I would love to be pointed their way."
    ),
    "H.IAAC": (
        "I am reaching out because the CAPIVARA/FairPIVARA line of work lives in a part of ML we care a lot about: multilingual multimodal data quality, evaluation design, and making model behavior legible rather than just better on aggregate. "
        "We are hiring for a role that sits very close to that territory for frontier labs. "
        "If there is a researcher or engineer on the team who led that work and might be open to a conversation, I would love an introduction."
    ),
}


@dataclass
class PinnedRepo:
    name: str
    description: str
    url: str


@dataclass
class RepoEvidence:
    name: str
    url: str
    exists: bool
    title: str
    meta_description: str
    readme_excerpt: str


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value)


@lru_cache(maxsize=None)
def load_text(path: str) -> str:
    file_path = Path(path)
    return file_path.read_text(encoding="utf-8") if file_path.exists() else ""


@lru_cache(maxsize=None)
def load_json_file(path: str) -> dict:
    file_path = Path(path)
    if not file_path.exists():
        return {}
    try:
        return json.loads(file_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def extract_username(github_url: str) -> str:
    path = re.sub(r"^https?://github\.com/", "", (github_url or "").strip())
    return path.strip("/").split("/")[0]


def clean_whitespace(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def strip_html(raw: str) -> str:
    raw = re.sub(r"(?is)<script.*?</script>", " ", raw)
    raw = re.sub(r"(?is)<style.*?</style>", " ", raw)
    raw = re.sub(r"(?is)<br\s*/?>", "\n", raw)
    raw = re.sub(r"(?is)</p>|</h\d>|</li>|</tr>|</section>|</article>", "\n", raw)
    raw = re.sub(r"(?is)<li[^>]*>", "- ", raw)
    raw = re.sub(r"(?is)<[^>]+>", " ", raw)
    raw = html.unescape(raw)
    return clean_whitespace(raw)


def extract_title(page_html: str) -> str:
    match = re.search(r"<title>(.*?)</title>", page_html, re.S | re.I)
    return clean_whitespace(html.unescape(match.group(1))) if match else ""


def extract_meta_description(page_html: str) -> str:
    match = re.search(
        r'<meta name="description" content="(.*?)"',
        page_html,
        re.S | re.I,
    )
    return clean_whitespace(html.unescape(match.group(1))) if match else ""


def normalize_profile_desc(text: str, username: str) -> str:
    text = clean_whitespace(text or "")
    text = text.replace(f" - {username}", "")
    text = re.sub(r"\b[^.]+ has \d+ repositories available\.?", "", text, flags=re.I)
    text = re.sub(r"Follow (their|his|her) code on GitHub\.?", "", text, flags=re.I)
    return clean_whitespace(text)


def extract_article_text(page_html: str) -> str:
    match = re.search(
        r'<article class="markdown-body entry-content container-lg" itemprop="text">(.*?)</article>',
        page_html,
        re.S | re.I,
    )
    if not match:
        match = re.search(
            r'<article class="markdown-body.*?itemprop="text">(.*?)</article>',
            page_html,
            re.S | re.I,
        )
    if not match:
        return ""
    return strip_html(match.group(1))


def detect_account_type(api: dict, profile_html: str) -> str:
    api_type = api.get("type")
    if api_type in {"User", "Organization"}:
        return api_type
    if 'content="organization:' in profile_html:
        return "Organization"
    if 'aria-label="Organization"' in profile_html:
        return "Organization"
    return "User"


def extract_pinned_repos(profile_html: str, username: str) -> list[PinnedRepo]:
    blocks = re.findall(
        r'<li\s+class="[^"]*pinned-item-list-item[^"]*"[^>]*>(.*?)</li>',
        profile_html,
        re.S | re.I,
    )
    repos: list[PinnedRepo] = []
    for block in blocks:
        name_match = re.search(r'<span class="repo">([^<]+)</span>', block)
        if not name_match:
            continue
        name = clean_whitespace(html.unescape(name_match.group(1)))
        desc_match = re.search(
            r'<p class="pinned-item-desc[^"]*">\s*(.*?)\s*</p>',
            block,
            re.S | re.I,
        )
        description = strip_html(desc_match.group(1)) if desc_match else ""
        repos.append(
            PinnedRepo(
                name=name,
                description=description,
                url=f"https://github.com/{username}/{name}",
            )
        )
    return repos


def extract_backticked_tokens(*values: str) -> list[str]:
    tokens: list[str] = []
    for value in values:
        tokens.extend(re.findall(r"`([^`]+)`", value or ""))
    return tokens


def extract_named_repo_tokens(*values: str) -> list[str]:
    patterns = [
        r"\b([A-Za-z0-9_.-]{3,})\s+repo\b",
        r"\b([A-Za-z0-9_.-]{3,})\s+library\b",
        r"\b([A-Za-z0-9_.-]{3,})\s+project\b",
        r"\b([A-Za-z0-9_.-]{3,})\s+benchmark\b",
        r"\b([A-Za-z0-9_.-]{3,})\s+toolkit\b",
        r"\b([A-Za-z0-9_.-]{3,})\s+work\b",
    ]
    tokens: list[str] = []
    for value in values:
        for pattern in patterns:
            tokens.extend(re.findall(pattern, value or "", re.I))
    return tokens


def dedupe_keep_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def repo_candidates(row_name: str, hook: str, notes: str, message: str) -> list[str]:
    raw = extract_backticked_tokens(hook, notes, message)
    raw.extend(extract_named_repo_tokens(hook, notes, message))
    clean: list[str] = []
    for token in raw:
        token = token.strip().strip(".,:;()[]{}")
        if not token:
            continue
        if "/" in token or " " in token:
            continue
        if token.lower() in STOPWORDS:
            continue
        if len(token) < 3:
            continue
        clean.append(token)
    return dedupe_keep_order(clean)[:4]


def fetch_repo_evidence(cache_dir: Path, username: str, repo_name: str) -> RepoEvidence:
    url = f"https://github.com/{username}/{repo_name}"
    page_html = load_text(str(cache_dir / f"repo__{safe_filename(username)}__{safe_filename(repo_name)}.html"))
    title = extract_title(page_html)
    meta_description = extract_meta_description(page_html)
    exists = "Page not found" not in title and bool(title)
    readme_text = extract_article_text(page_html)
    excerpt = readme_text[:700].strip()
    return RepoEvidence(
        name=repo_name,
        url=url,
        exists=exists,
        title=title,
        meta_description=meta_description,
        readme_excerpt=excerpt,
    )


def is_generic_repo_description(text: str) -> bool:
    text = (text or "").strip()
    if not text:
        return True
    generic_prefixes = [
        "Contribute to ",
        "GitHub - ",
    ]
    if any(text.startswith(prefix) for prefix in generic_prefixes):
        return True
    return text.endswith("creating an account on GitHub.")


def first_meaningful_readme_line(text: str) -> str:
    for raw_line in text.splitlines():
        line = clean_whitespace(raw_line.strip("#- "))
        if len(line) < 12:
            continue
        if re.match(r"^(running|installation|usage|authors|paper link|implementation)$", line, re.I):
            continue
        return line
    return clean_whitespace(text.split(".")[0]) if text else ""


def best_repo_summary(repo: RepoEvidence) -> str:
    if repo.meta_description and not is_generic_repo_description(repo.meta_description):
        return repo.meta_description
    readme_line = first_meaningful_readme_line(repo.readme_excerpt)
    if readme_line and not is_generic_repo_description(readme_line):
        return readme_line
    return repo.name


def choose_primary_evidence(
    cache_dir: Path,
    username: str,
    pinned_repos: list[PinnedRepo],
    repo_names: list[str],
) -> tuple[list[RepoEvidence], bool]:
    evidence: list[RepoEvidence] = []
    pinned_lookup = {repo.name.lower(): repo for repo in pinned_repos}
    for repo_name in repo_names:
        matched = pinned_lookup.get(repo_name.lower())
        target_name = matched.name if matched else repo_name
        repo = fetch_repo_evidence(cache_dir, username, target_name)
        if repo.exists:
            evidence.append(repo)
    if evidence:
        return evidence[:2], False
    if repo_names:
        fallback: list[RepoEvidence] = []
        for pinned in pinned_repos[:2]:
            repo = fetch_repo_evidence(cache_dir, username, pinned.name)
            if repo.exists:
                fallback.append(repo)
        return fallback[:2], True
    for pinned in pinned_repos[:2]:
        repo = fetch_repo_evidence(cache_dir, username, pinned.name)
        if repo.exists:
            evidence.append(repo)
    return evidence[:2], False


def summarize_primary_source(
    profile_desc: str,
    profile_readme: str,
    pinned_repos: list[PinnedRepo],
    repo_evidence: list[RepoEvidence],
    fallback_used: bool,
    had_repo_candidates: bool,
) -> str:
    parts: list[str] = []
    if fallback_used and had_repo_candidates:
        parts.append("The repo named in the current outreach copy was not strongly corroborated from the public GitHub surface.")
        if repo_evidence:
            primary = repo_evidence[0]
            parts.append(f"Most visible public repo is `{primary.name}`.")
            parts.append(best_repo_summary(primary))
        elif profile_desc:
            parts.append(profile_desc)
    elif repo_evidence:
        primary = repo_evidence[0]
        parts.append(f"`{primary.name}` exists publicly on the profile.")
        parts.append(best_repo_summary(primary))
    elif pinned_repos:
        pinned = pinned_repos[0]
        parts.append(f"`{pinned.name}` is pinned on the profile.")
        if pinned.description:
            parts.append(pinned.description)
    elif profile_desc:
        parts.append(profile_desc)
    elif profile_readme:
        parts.append(profile_readme[:220].rstrip() + "...")
    return clean_whitespace(" ".join(parts))


def build_safer_hook(
    row_name: str,
    profile_desc: str,
    pinned_repos: list[PinnedRepo],
    repo_evidence: list[RepoEvidence],
    fallback_used: bool,
    had_repo_candidates: bool,
) -> str:
    if fallback_used and had_repo_candidates:
        if profile_desc:
            return profile_desc[:140].rstrip()
        return f"{row_name}'s broader ML/AI work on GitHub"
    if repo_evidence:
        primary = repo_evidence[0]
        text = best_repo_summary(primary)
        text = clean_whitespace(text).rstrip(".")
        if len(text) > 140:
            text = text[:137].rstrip() + "..."
        return text
    if pinned_repos:
        text = f"{pinned_repos[0].name}: {pinned_repos[0].description}".strip(": ")
        return text[:140].rstrip()
    if profile_desc:
        return profile_desc[:140].rstrip()
    return f"public GitHub work from {row_name}"


def build_safer_notes(
    username: str,
    profile_desc: str,
    profile_readme: str,
    pinned_repos: list[PinnedRepo],
    repo_evidence: list[RepoEvidence],
    account_type: str,
    fallback_used: bool,
    had_repo_candidates: bool,
) -> str:
    notes: list[str] = []
    if account_type.lower() == "organization":
        notes.append("This GitHub URL is an organization account, not an individual profile.")
    if profile_desc:
        notes.append(f"Profile bio: {profile_desc}")
    if fallback_used and had_repo_candidates:
        notes.append(
            "The repo named in the current message was not strongly corroborated from the public GitHub surface, so the outreach should stay broad."
        )
        if repo_evidence:
            notes.append(f"The most visible public repo is `{repo_evidence[0].name}`, but it should not replace the original claim one-for-one.")
        return clean_whitespace(" ".join(notes))
    if repo_evidence:
        primary = repo_evidence[0]
        notes.append(f"Best public source is `{primary.name}`.")
        notes.append(f"Public repo surface supports the hook: {best_repo_summary(primary)}")
    elif pinned_repos:
        notes.append(
            f"Pinned repo surface is thin; `{pinned_repos[0].name}` is visible but the detailed claim should stay conservative."
        )
    elif profile_readme:
        notes.append("Most support comes from the profile README rather than a repo README.")
    else:
        notes.append("Public GitHub surface is thin; keep claims high-level and avoid over-specific technical assertions.")
    return clean_whitespace(" ".join(notes))


def classify_verdict(
    account_type: str,
    original_hook: str,
    original_notes: str,
    had_repo_candidates: bool,
    fallback_used: bool,
    repo_evidence: list[RepoEvidence],
    pinned_repos: list[PinnedRepo],
    profile_desc: str,
) -> tuple[str, str]:
    if account_type.lower() == "organization":
        return ("ORG_ACCOUNT", "high")
    if had_repo_candidates and not repo_evidence:
        return ("THIN", "high")
    if had_repo_candidates and fallback_used:
        return ("CAUTION", "high")
    if repo_evidence and (original_hook or original_notes):
        return ("SAFE", "low")
    if repo_evidence:
        return ("CAUTION", "medium")
    if pinned_repos or profile_desc:
        return ("THIN", "medium")
    return ("THIN", "high")


def first_name(name: str) -> str:
    clean = (name or "").strip()
    if not clean:
        return "there"
    return clean.split()[0]


def shorten_for_subject(text: str, fallback: str) -> str:
    text = re.sub(r"https?://\S+", "", text or "")
    text = re.sub(r"`", "", text)
    text = clean_whitespace(text).rstrip(".")
    if not text:
        return fallback
    if len(text) > 55:
        text = text[:52].rstrip() + "..."
    return text


def shorten_for_message(text: str, fallback: str) -> str:
    text = re.sub(r"\s+-\s+[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+$", "", text or "")
    text = clean_whitespace(text).rstrip(".")
    if not text:
        return fallback
    if len(text) > 95:
        text = text[:92].rstrip() + "..."
    return text


def build_subject(name: str, safer_hook: str, account_type: str) -> str:
    if account_type.lower() == "organization":
        return f"{shorten_for_subject(safer_hook, name)} and a frontier data role"
    return f"Your {shorten_for_subject(safer_hook, 'GitHub work')} and a frontier data role"


def build_streamlined_message(
    name: str,
    safer_hook: str,
    account_type: str,
) -> str:
    hook_text = shorten_for_message(safer_hook, "your GitHub work")
    hook_clause = hook_text if hook_text.endswith("...") else f"{hook_text}."
    if account_type.lower() == "organization":
        return (
            f"Hi team, I came across your GitHub work on {hook_clause} "
            f"We're hiring for {ROLE_PITCH}, and this looked relevant. "
            "If there's a researcher or engineer on the team who would be open to hearing about it, I'd love to connect."
        )
    return (
        f"Hi {first_name(name)}, I came across your GitHub work on {hook_clause} "
        f"We're hiring for {ROLE_PITCH}, and your background felt relevant. "
        "If you're open to it, I'd be happy to share a bit more context."
    )


def build_elevated_subject(
    name: str,
    original_subject: str,
    revised_subject: str,
    verdict: str,
    risk: str,
) -> str:
    if name in ELEVATED_SUBJECT_OVERRIDES:
        return ELEVATED_SUBJECT_OVERRIDES[name]
    if original_subject and (verdict == "SAFE" or (verdict == "CAUTION" and risk != "high")):
        return original_subject
    return original_subject or revised_subject


def build_elevated_message(
    name: str,
    revised_message: str,
) -> str:
    return ELEVATED_MESSAGE_OVERRIDES.get(name, revised_message)


def primary_sources(username: str, repo_evidence: list[RepoEvidence]) -> str:
    urls = [f"https://github.com/{username}"]
    urls.extend(repo.url for repo in repo_evidence)
    return "\n".join(dedupe_keep_order(urls))


def set_wrap(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--out-xlsx", required=True)
    parser.add_argument("--out-md", required=True)
    parser.add_argument("--cache-dir", required=True)
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    out_xlsx = Path(args.out_xlsx)
    out_md = Path(args.out_md)
    cache_dir = Path(args.cache_dir)

    wb = load_workbook(workbook_path)
    ws = wb["GitHub Outreach"]

    headers = {
        12: "Audit Verdict",
        13: "Audit Risk",
        14: "Evidence Summary",
        15: "Revised Subject",
        16: "Revised Message",
        17: "Revised Hook",
        18: "Revised Notes",
        19: "Primary Sources",
        20: "Elevated Subject",
        21: "Elevated Message",
    }
    for col, label in headers.items():
        ws.cell(row=1, column=col).value = label

    summary_lines = [
        "# GitHub Outreach Audit",
        "",
        f"Source workbook: `{workbook_path}`",
        "",
    ]

    counts = {"SAFE": 0, "CAUTION": 0, "THIN": 0, "ORG_ACCOUNT": 0}

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value or ""
        github_url = ws.cell(row=row, column=4).value or ""
        subject = ws.cell(row=row, column=8).value or ""
        message = ws.cell(row=row, column=9).value or ""
        hook = ws.cell(row=row, column=10).value or ""
        notes = ws.cell(row=row, column=11).value or ""

        if not github_url:
            continue

        username = extract_username(github_url)
        api = load_json_file(str(cache_dir / f"user__{safe_filename(username)}.json"))
        profile_html = load_text(str(cache_dir / f"profile__{safe_filename(username)}.html"))
        account_type = detect_account_type(api, profile_html)
        profile_desc = normalize_profile_desc(extract_meta_description(profile_html), username)
        profile_readme = extract_article_text(profile_html)
        pinned_repos = extract_pinned_repos(profile_html, username)
        candidates = repo_candidates(name, hook, notes, message)
        repo_evidence, fallback_used = choose_primary_evidence(cache_dir, username, pinned_repos, candidates)

        safer_hook = build_safer_hook(
            name,
            profile_desc,
            pinned_repos,
            repo_evidence,
            fallback_used=fallback_used,
            had_repo_candidates=bool(candidates),
        )
        safer_notes = build_safer_notes(
            username=username,
            profile_desc=profile_desc,
            profile_readme=profile_readme,
            pinned_repos=pinned_repos,
            repo_evidence=repo_evidence,
            account_type=account_type,
            fallback_used=fallback_used,
            had_repo_candidates=bool(candidates),
        )
        verdict, risk = classify_verdict(
            account_type=account_type,
            original_hook=hook,
            original_notes=notes,
            had_repo_candidates=bool(candidates),
            fallback_used=fallback_used,
            repo_evidence=repo_evidence,
            pinned_repos=pinned_repos,
            profile_desc=profile_desc,
        )
        counts[verdict] += 1

        evidence_summary = summarize_primary_source(
            profile_desc=profile_desc,
            profile_readme=profile_readme,
            pinned_repos=pinned_repos,
            repo_evidence=repo_evidence,
            fallback_used=fallback_used,
            had_repo_candidates=bool(candidates),
        )
        revised_subject = build_subject(name, safer_hook, account_type)
        revised_message = build_streamlined_message(name, safer_hook, account_type)
        elevated_subject = build_elevated_subject(
            name=name,
            original_subject=subject,
            revised_subject=revised_subject,
            verdict=verdict,
            risk=risk,
        )
        elevated_message = build_elevated_message(name, revised_message)
        sources = primary_sources(username, repo_evidence)

        ws.cell(row=row, column=12).value = verdict
        ws.cell(row=row, column=13).value = risk
        ws.cell(row=row, column=14).value = evidence_summary
        ws.cell(row=row, column=15).value = revised_subject
        ws.cell(row=row, column=16).value = revised_message
        ws.cell(row=row, column=17).value = safer_hook
        ws.cell(row=row, column=18).value = safer_notes
        ws.cell(row=row, column=19).value = sources
        ws.cell(row=row, column=20).value = elevated_subject
        ws.cell(row=row, column=21).value = elevated_message
        set_wrap(ws, row, 14, 21)

        summary_lines.extend(
            [
                f"## {name}",
                "",
                f"- GitHub: {github_url}",
                f"- Verdict: {verdict}",
                f"- Risk: {risk}",
                f"- Original Subject: {subject}",
                f"- Evidence Summary: {evidence_summary}",
                f"- Revised Hook: {safer_hook}",
                f"- Revised Notes: {safer_notes}",
                f"- Revised Message: {revised_message}",
                f"- Elevated Subject: {elevated_subject}",
                f"- Elevated Message: {elevated_message}",
                f"- Sources: {sources.replace(chr(10), ', ')}",
                "",
            ]
        )

    ws.freeze_panes = "A2"
    ws.column_dimensions["N"].width = 55
    ws.column_dimensions["O"].width = 36
    ws.column_dimensions["P"].width = 70
    ws.column_dimensions["Q"].width = 40
    ws.column_dimensions["R"].width = 70
    ws.column_dimensions["S"].width = 45
    ws.column_dimensions["T"].width = 42
    ws.column_dimensions["U"].width = 90

    wb.save(out_xlsx)

    summary_lines[4:4] = [
        "## Counts",
        "",
        f"- SAFE: {counts['SAFE']}",
        f"- CAUTION: {counts['CAUTION']}",
        f"- THIN: {counts['THIN']}",
        f"- ORG_ACCOUNT: {counts['ORG_ACCOUNT']}",
        "",
    ]
    out_md.write_text("\n".join(summary_lines), encoding="utf-8")


if __name__ == "__main__":
    main()
