#!/usr/bin/env python3
"""Export canonical LinkedIn save candidates into a review workbook.

This uses runtime_state.sqlite3 as the source of truth and stamps rows into an
existing review workbook template so the resulting XLSX preserves the exact
layout, column widths, pane freezing, and cell formatting expected by recruiters.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


SAVE_DECISIONS = ("SAVE", "INFERENTIAL_SAVE", "TRANSFERABLE_SAVE", "SIGNAL_SAVE")
DEFAULT_PIPELINE_SUFFIX = " (uncontacted)"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", required=True, help="Path to runtime_state.sqlite3")
    parser.add_argument("--template", required=True, help="Path to reference XLSX template")
    parser.add_argument("--out", required=True, help="Path to output XLSX")
    parser.add_argument(
        "--pipeline-status",
        default="",
        help="Value to write into the Pipeline Status column for each candidate",
    )
    return parser.parse_args()


def _connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _first_nonempty(*values: object) -> str:
    for value in values:
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _join_chunks(chunks: list[str], *, limit: int | None = None) -> str:
    filtered = [chunk.strip() for chunk in chunks if chunk and str(chunk).strip()]
    if limit is not None:
        filtered = filtered[:limit]
    return "; ".join(filtered)


def _format_experience(experiences: list[dict], *, limit: int = 3) -> str:
    chunks: list[str] = []
    for exp in experiences[:limit]:
        title = (exp.get("title") or "").strip()
        company = (exp.get("company") or "").strip()
        if title and company:
            chunks.append(f"{title} @ {company}")
        elif title or company:
            chunks.append(title or company)
    return _join_chunks(chunks)


def _format_education(entries: list[dict], *, limit: int = 4) -> str:
    chunks: list[str] = []
    for entry in entries[:limit]:
        degree = (entry.get("degree") or "").strip()
        school = (entry.get("school") or "").strip()
        field = (entry.get("field") or "").strip()
        if degree and school:
            chunks.append(f"{degree}, {school}")
        elif school and field:
            chunks.append(f"{school}, {field}")
        elif degree or school or field:
            chunks.append(degree or school or field)
    return _join_chunks(chunks)


def _signal_and_confidence(confidence: float | None) -> tuple[str, str]:
    if confidence is None:
        return "", ""
    pct = round(float(confidence) * 100)
    if pct >= 85:
        return "Very Strong", f"{pct}%"
    if pct >= 72:
        return "Strong", f"{pct}%"
    if pct >= 55:
        return "Moderate", f"{pct}%"
    return "", ""


def _assessment_rationale(full_decision: dict) -> str:
    rationale = (full_decision.get("rationale") or "").strip()
    modifier = (full_decision.get("post_save_modifier") or "").strip()
    if modifier and modifier.upper() != "NONE":
        return f"{rationale} | {modifier}" if rationale else modifier
    return rationale


def _pipeline_status(conn: sqlite3.Connection, fallback: str) -> str:
    if fallback.strip():
        return fallback.strip()
    row = conn.execute(
        """
        SELECT json_extract(resume_state_json, '$.linkedin_project') AS linkedin_project
        FROM runs
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    project = (row["linkedin_project"] or "").strip() if row else ""
    return f"{project}{DEFAULT_PIPELINE_SUFFIX}" if project else ""


def _load_candidates(conn: sqlite3.Connection) -> list[dict]:
    rows = conn.execute(
        f"""
        SELECT display_name, profile_url, terminal_decision, terminal_payload_json
        FROM candidates
        WHERE terminal_decision IN ({",".join("?" for _ in SAVE_DECISIONS)})
        ORDER BY json_extract(terminal_payload_json, '$.full_decision.confidence') DESC,
                 display_name COLLATE NOCASE
        """,
        SAVE_DECISIONS,
    ).fetchall()

    candidates: list[dict] = []
    for row in rows:
        payload = json.loads(row["terminal_payload_json"])
        full_decision = payload.get("full_decision", {})
        profile_summary = payload.get("profile_summary", {})
        snippet = payload.get("snippet", {})
        experiences = profile_summary.get("experiences") or []
        education = profile_summary.get("education") or []
        top_experience = experiences[0] if experiences else {}

        signal_strength, confidence_text = _signal_and_confidence(full_decision.get("confidence"))
        assessment_path = (full_decision.get("path") or "").strip()
        if assessment_path.lower() == "none":
            assessment_path = ""

        candidates.append(
            {
                "Name": (row["display_name"] or "").strip(),
                "Signal Strength": signal_strength,
                "Confidence": confidence_text,
                "Current Title": _first_nonempty(top_experience.get("title"), snippet.get("current_title")),
                "Company": _first_nonempty(top_experience.get("company"), snippet.get("current_company")),
                "Location": _first_nonempty(top_experience.get("location"), snippet.get("location")),
                "Headline": _first_nonempty(profile_summary.get("headline"), snippet.get("headline")),
                "Key Experience": _format_experience(experiences) or _join_chunks(snippet.get("experience_entries") or [], limit=3),
                "Education": _format_education(education) or (snippet.get("education_snippet") or "").strip(),
                "Assessment Path": assessment_path,
                "Assessment Rationale": _assessment_rationale(full_decision),
                "LinkedIn": "Profile",
                "LinkedIn URL": (row["profile_url"] or "").strip(),
                "Pipeline Status": "",
                "OUTREACH": "",
                "Recruiter Review Decision": "",
                "Feedback": "",
                "_confidence_numeric": float(full_decision.get("confidence") or 0.0),
                "_terminal_decision": row["terminal_decision"],
            }
        )

    return candidates


def _style_cell_from(target, source) -> None:
    target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format


def _prepare_workbook(template_path: Path):
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    return wb, ws


def _capture_style_templates(ws) -> dict[str, object]:
    return {
        "row_even": [copy(ws.cell(34, col_idx)._style) for col_idx in range(1, 19)],
        "row_odd": [copy(ws.cell(33, col_idx)._style) for col_idx in range(1, 19)],
        "signal_very_strong": copy(ws.cell(2, 2)._style),
        "signal_strong": copy(ws.cell(5, 2)._style),
        "signal_moderate": copy(ws.cell(21, 2)._style),
        "blank_review_16": copy(ws.cell(37, 16)._style),
        "blank_review_17": copy(ws.cell(37, 17)._style),
        "blank_review_18": copy(ws.cell(37, 18)._style),
    }


def _clear_existing_rows(ws) -> None:
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, 19):
            cell = ws.cell(row_idx, col_idx)
            cell.value = None
            cell.hyperlink = None


def _apply_row_styles(ws, row_idx: int, *, signal_strength: str, style_templates: dict[str, object]) -> None:
    row_styles = style_templates["row_even"] if (row_idx - 2) % 2 == 0 else style_templates["row_odd"]
    for col_idx in range(1, 19):
        ws.cell(row_idx, col_idx)._style = copy(row_styles[col_idx - 1])

    ws.cell(row_idx, 16)._style = copy(style_templates["blank_review_16"])
    ws.cell(row_idx, 17)._style = copy(style_templates["blank_review_17"])
    ws.cell(row_idx, 18)._style = copy(style_templates["blank_review_18"])

    signal_style = {
        "Very Strong": style_templates["signal_very_strong"],
        "Strong": style_templates["signal_strong"],
        "Moderate": style_templates["signal_moderate"],
    }.get(signal_strength)
    if signal_style is not None:
        ws.cell(row_idx, 2)._style = copy(signal_style)


def _write_candidates(ws, candidates: list[dict], pipeline_status: str, style_templates: dict[str, object]) -> None:
    for idx, candidate in enumerate(candidates, start=2):
        _apply_row_styles(
            ws,
            idx,
            signal_strength=candidate["Signal Strength"],
            style_templates=style_templates,
        )

        ws.cell(idx, 1).value = candidate["Name"]
        ws.cell(idx, 2).value = candidate["Signal Strength"] or None
        ws.cell(idx, 3).value = candidate["Confidence"] or None
        ws.cell(idx, 4).value = candidate["Current Title"] or None
        ws.cell(idx, 5).value = candidate["Company"] or None
        ws.cell(idx, 6).value = candidate["Location"] or None
        ws.cell(idx, 7).value = candidate["Headline"] or None
        ws.cell(idx, 8).value = candidate["Key Experience"] or None
        ws.cell(idx, 9).value = candidate["Education"] or None
        ws.cell(idx, 10).value = candidate["Assessment Path"] or None
        ws.cell(idx, 11).value = candidate["Assessment Rationale"] or None
        ws.cell(idx, 12).value = "Profile"
        ws.cell(idx, 12).hyperlink = candidate["LinkedIn URL"]
        ws.cell(idx, 13).value = pipeline_status or None
        ws.cell(idx, 16).value = None
        ws.cell(idx, 17).value = None
        ws.cell(idx, 18).value = None

    filter_end = max(2, len(candidates) + 1)
    ws.auto_filter.ref = f"$A$1:$M${filter_end}"


def main() -> None:
    args = _parse_args()
    db_path = Path(args.db)
    template_path = Path(args.template)
    out_path = Path(args.out)

    conn = _connect(db_path)
    try:
        candidates = _load_candidates(conn)
        pipeline_status = _pipeline_status(conn, args.pipeline_status)
    finally:
        conn.close()

    wb, ws = _prepare_workbook(template_path)
    style_templates = _capture_style_templates(ws)
    _clear_existing_rows(ws)
    _write_candidates(ws, candidates, pipeline_status, style_templates)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {len(candidates)} candidates to {out_path}")


if __name__ == "__main__":
    main()
